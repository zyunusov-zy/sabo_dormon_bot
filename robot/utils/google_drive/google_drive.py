import os
import io
from dotenv import load_dotenv
from google.oauth2.service_account import Credentials as ServiceAccountCredentials
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseUpload
from aiogram import Bot
from googleapiclient.errors import HttpError
import pandas as pd
from robot.utils.misc.logging import log_handler, log_user_action, log_state_change, log_error, log_file_operation

load_dotenv(dotenv_path=".env", override=True)

SCOPES = ['https://www.googleapis.com/auth/drive.file']
SERVICE_ACCOUNT_PATH = os.getenv('GOOGLE_DRIVE_SERVICE_ACCOUNT_PATH')
PARENT_FOLDER_ID = os.getenv("GOOGLE_DRIVE_PARENT_FOLDER_ID")

QUESTION_LABELS = {
    "q1_full_name": "ФИО",
    "q2_birth_date": "Дата рождения",
    "q3_gender": "Пол",
    "q4_phone_number": "Телефон",
    "q5_telegram_username": "Telegram",
    "q6_region": "Регион",
    "q7_who_applies": "Кто обращается за помощью",
    "q8_is_sabodarmon": "Пациент Sabo?",
    "q9_source_info": "Как узнали о нас?",
    "q10_has_diagnosis": "Есть ли диагноз?",
    "q11_diagnosis_text": "Укажите диагноз",
    "q13_complaint": "Какие жалобы?",
    "q14_main_discomfort": "Что мешает больше всего?",
    "q15_improvements": "Что улучшится после лечения?",
    "q16_consequences": "Что будет, если не лечиться?",
    "q17_need_confirmation": "Нужны подтверждающие документы?",
    "q18_avg_income": "Средний доход на члена семьи (в млн сум)",
    "q19_children_count": "Сколько детей в семье?",
    "q21_family_work": "Кто работает в семье?",
    "q22_housing_type": "Какое у вас жильё?",
    "q23_diagnosis_confirm": "Подтверждение диагноза",
    "q25_final_comment": "Комментарий к заявке",

    # Файлы
    "q12_diagnosis_file_id": "📎 Диагноз (файл)",
    "q17_confirmation_file": "📎 Подтверждающие документы (файл)",
    "q18_income_doc": "📎 Справка о доходах (файл)",
    "q19_children_docs": "📎 Документы на детей (файлы)",
    "q22_housing_doc": "📎 Документ на жильё (файл)",
    "q24_additional_file": "📎 Дополнительный файл",
}

QUESTION_FILE_KEYS = {
    "q12_diagnosis_file_id": "📎 Диагноз (файл)",
    "q17_confirmation_file": "📎 Подтверждающие документы (файл)",
    "q18_income_doc": "📎 Справка о доходах (файл)",
    "q19_children_docs": "📎 Документы на детей (файлы)",
    "q22_housing_doc": "📎 Документ на жильё (файл)",
    "q24_additional_file": "📎 Дополнительный файл",
}

def get_drive_service():
    if not SERVICE_ACCOUNT_PATH:
        raise ValueError("❌ GOOGLE_DRIVE_SERVICE_ACCOUNT_PATH not set in .env file")
    
    if not os.path.exists(SERVICE_ACCOUNT_PATH):
        raise FileNotFoundError(f"❌ Service account file not found: {SERVICE_ACCOUNT_PATH}")
    
    creds = ServiceAccountCredentials.from_service_account_file(SERVICE_ACCOUNT_PATH, scopes=SCOPES)
    return build('drive', 'v3', credentials=creds)

def create_folder(name: str, parent_id: str = None) -> str:
    service = get_drive_service()
    metadata = {'name': name, 'mimeType': 'application/vnd.google-apps.folder'}
    if parent_id:
        metadata['parents'] = [parent_id]
    folder = service.files().create(body=metadata, fields='id').execute()
    return folder['id']

def upload_file_to_folder(file_bytes: bytes, filename: str, mime_type: str, folder_id: str) -> str:
    service = get_drive_service()
    metadata = {'name': filename, 'parents': [folder_id]}
    media = MediaIoBaseUpload(io.BytesIO(file_bytes), mimetype=mime_type)
    uploaded_file = service.files().create(body=metadata, media_body=media, fields='id').execute()
    return uploaded_file['id']

def create_questionnaire_excel_bytes(user_data: dict) -> bytes:
    """Создает Excel файл анкеты в памяти и возвращает байты"""
    try:
        # Подготавливаем данные для таблицы
        questions = []
        answers = []
        
        # Добавляем ФИО первым
        full_name = user_data.get('q1_full_name', '')
        if full_name:
            questions.append("ФИО")
            answers.append(full_name)

        # Обрабатываем остальные вопросы
        for key, value in user_data.items():
            if key == "q1_full_name" or key == "full_name":
                continue
                
            label = QUESTION_LABELS.get(key, key)

            if key in QUESTION_FILE_KEYS:
                # Для файлов показываем статус наличия
                if isinstance(value, list):
                    answer = f"{len(value)} файл(ов) ✅" if value else "—"
                else:
                    answer = "Есть файл ✅" if value else "—"
            else:
                # Для обычных ответов
                answer = str(value) if value is not None else "—"
            
            questions.append(label)
            answers.append(answer)

        # Создаем DataFrame
        df = pd.DataFrame({
            'Вопрос': questions,
            'Ответ': answers
        })

        # Создаем Excel в памяти
        excel_buffer = io.BytesIO()
        
        # Создаем Excel writer с настройками
        with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Анкета', index=False)
            
            # Получаем worksheet для настройки форматирования
            worksheet = writer.sheets['Анкета']
            
            # Настраиваем ширину столбцов
            worksheet.column_dimensions['A'].width = 45  # Столбец "Вопрос"
            worksheet.column_dimensions['B'].width = 65  # Столбец "Ответ"
            
            # Импортируем стили для форматирования
            from openpyxl.styles import Alignment, Border, Side, PatternFill, Font
            
            # Определяем стили
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")  # Синий фон для заголовков
            header_font = Font(color="FFFFFF", bold=True, size=12)  # Белый жирный текст
            
            question_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")  # Светло-серый для вопросов
            question_font = Font(bold=True, size=11)  # Жирный текст для вопросов
            
            answer_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")  # Белый для ответов
            answer_font = Font(size=11)  # Обычный текст для ответов
            
            # Создаем границы
            thin_border = Border(
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000'),
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000')
            )
            
            # Форматируем заголовки (первая строка)
            for col in range(1, 3):  # Столбцы A и B
                cell = worksheet.cell(row=1, column=col)
                cell.fill = header_fill
                cell.font = header_font
                cell.border = thin_border
                cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
            
            # Форматируем остальные строки
            for row in range(2, len(df) + 2):  # Начинаем со второй строки
                # Столбец A (Вопросы) - серый фон
                question_cell = worksheet.cell(row=row, column=1)
                question_cell.fill = question_fill
                question_cell.font = question_font
                question_cell.border = thin_border
                question_cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')
                
                # Столбец B (Ответы) - белый фон
                answer_cell = worksheet.cell(row=row, column=2)
                answer_cell.fill = answer_fill
                answer_cell.font = answer_font
                answer_cell.border = thin_border
                answer_cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')
            
            # Устанавливаем высоту строк для лучшего отображения
            for row in range(1, len(df) + 2):
                worksheet.row_dimensions[row].height = 25

        # Получаем байты
        excel_buffer.seek(0)
        return excel_buffer.read()

    except Exception as e:
        print(f"❌ Ошибка при создании Excel: {e}")
        raise

def upload_excel_to_drive(excel_bytes: bytes, filename: str, folder_id: str) -> str:
    """Загружает Excel файл на Google Drive"""
    service = get_drive_service()
    metadata = {
        'name': filename,
        'parents': [folder_id]
    }
    media = MediaIoBaseUpload(
        io.BytesIO(excel_bytes), 
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    uploaded_file = service.files().create(body=metadata, media_body=media, fields='id').execute()
    return uploaded_file['id']

def upload_text_to_drive(content: str, filename: str, folder_id: str = None) -> str:
    """Загружает текстовый файл на Google Drive"""
    service = get_drive_service()
    metadata = {'name': filename, 'mimeType': 'text/plain'}
    if folder_id:
        metadata['parents'] = [folder_id]
    media = MediaIoBaseUpload(io.BytesIO(content.encode('utf-8')), mimetype='text/plain')
    file = service.files().create(body=metadata, media_body=media, fields='id').execute()
    return file['id']

async def save_file_by_id(file_id: str, folder_id: str, filename: str, bot: Bot, user_id: int = None):
    try:
        file = await bot.get_file(file_id)
        file_path = file.file_path
        file_bytes = await bot.download_file(file_path)
        ext = os.path.splitext(file_path)[-1] or ".bin"

        mime_type = "application/octet-stream"
        if ext.lower() in ['.jpg', '.jpeg']:
            mime_type = "image/jpeg"
        elif ext.lower() == '.png':
            mime_type = "image/png"
        elif ext.lower() == '.pdf':
            mime_type = "application/pdf"
        elif ext.lower() in ['.doc', '.docx']:
            mime_type = "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
        elif ext.lower() == '.txt':
            mime_type = "text/plain"

        log_user_action(
            user_id=user_id,
            action="Uploading file",
            state="save_file_by_id",
            extra_data=f"Filename: {filename}{ext}, FolderID: {folder_id}"
        )

        return upload_file_to_folder(file_bytes.read(), f"{filename}{ext}", mime_type, folder_id)

    except Exception as e:
        log_error(
            user_id=user_id,
            error=e,
            action="Failed to upload file",
            state="save_file_by_id",
        )
        print(f"Error saving file {file_id}: {e}")
        raise

async def save_full_questionnaire_to_drive(user_data: dict, bot: Bot, folder_id: str, user_id: int = None):
    try:
        full_name = user_data.get('q1_full_name', 'Пациент')
        birth_date = user_data.get('q2_birth_date', 'Unknown')
        root_folder_id = folder_id

        # Создаем Excel файл анкеты
        excel_bytes = create_questionnaire_excel_bytes(user_data)
        
        log_user_action(
            user_id=user_id,
            action="Uploading questionnaire Excel",
            state="save_full_questionnaire_to_drive",
            extra_data="Анкета.xlsx"
        )

        # Загружаем Excel файл на Drive
        upload_excel_to_drive(excel_bytes, "Анкета.xlsx", folder_id=root_folder_id)

        # Создаем папку для файлов
        files_folder_id = create_folder("Файлы", parent_id=root_folder_id)
        
        log_user_action(
            user_id=user_id,
            action="Created files folder for patient",
            state="save_full_questionnaire_to_drive",
            extra_data=f"Files folder ID: {files_folder_id}"
        )

        # Загружаем вложения в подпапки внутри папки "Файлы"
        for field, folder_name in QUESTION_FILE_KEYS.items():
            file_value = user_data.get(field)
            if not file_value:
                continue

            # Создаем подпапку для каждого типа документов внутри папки "Файлы"
            subfolder_id = create_folder(folder_name, parent_id=files_folder_id)

            log_user_action(
                user_id=user_id,
                action="Created subfolder for attachments",
                state="save_full_questionnaire_to_drive",
                extra_data=f"{folder_name} (Field: {field})"
            )

            if isinstance(file_value, list):
                # Если несколько файлов
                for idx, file_id in enumerate(file_value):
                    if file_id:
                        await save_file_by_id(file_id, subfolder_id, f"{folder_name}_{idx+1}", bot, user_id)
            else:
                # Если один файл
                await save_file_by_id(file_value, subfolder_id, folder_name, bot, user_id)

        log_user_action(
            user_id=user_id,
            action="Finished saving questionnaire",
            state="save_full_questionnaire_to_drive"
        )

    except Exception as e:
        log_error(
            user_id=user_id,
            error=e,
            action="Error saving questionnaire",
            state="save_full_questionnaire_to_drive",
        )
        print(f"❌ Error in save_full_questionnaire_to_drive: {e}")
        raise

def delete_folder(folder_id: str):
    try:
        service = get_drive_service()
        service.files().delete(fileId=folder_id).execute()
        print(f"✅ Папка с ID {folder_id} успешно удалена.")
        return True
    except HttpError as e:
        print(f"❌ Ошибка при удалении папки: {e}")
        return False