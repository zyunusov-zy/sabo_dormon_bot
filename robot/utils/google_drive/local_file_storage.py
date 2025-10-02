import os
import asyncio
from pathlib import Path
from aiogram import Bot
import pandas as pd
from robot.utils.misc.logging import log_handler, log_user_action, log_state_change, log_error, log_file_operation

# Базовая папка для сохранения всех анкет
BASE_STORAGE_PATH = "questionnaire_storage"

QUESTION_LABELS = {
    "q1_full_name": "ФИО",
    "q2_birth_date": "Дата рождения",
    "q3_gender": "Пол",
    "q4_phone_number": "Телефон",
    "q5_telegram_username": "Telegram",
    "q6_region": "Регион",
    "q7_who_applies": "Кто обращается за помощью",
    "q8_is_sabodarmon": "Пациент Sabo?",
    "q9_source_info": "Как узнали о нас?",
    "q10_has_diagnosis": "Есть ли диагноз?",
    "q11_diagnosis_text": "Укажите диагноз",
    "q13_complaint": "Какие жалобы?",
    "q14_main_discomfort": "Что мешает больше всего?",
    "q15_improvements": "Что улучшится после лечения?",
    "q16_consequences": "Что будет, если не лечиться?",
    "q17_need_confirmation": "Нужны подтверждающие документы?",
    "q18_avg_income": "Средний доход на члена семьи (в млн сум)",
    "q19_children_count": "Сколько детей в семье?",
    "q21_family_work": "Кто работает в семье?",
    "q22_housing_type": "Какое у вас жильё?",
    "q23_diagnosis_confirm": "Подтверждение диагноза",
    "q25_final_comment": "Комментарий к заявке",

    # Файлы
    "q12_diagnosis_file_id": "📎 Диагноз (файл)",
    "q17_confirmation_file": "📎 Подтверждающие документы (файл)",
    "q18_income_doc": "📎 Справка о доходах (файл)",
    "q19_children_docs": "📎 Документы на детей (файлы)",
    "q22_housing_doc": "📎 Документ на жильё (файл)",
    "q24_additional_file": "📎 Дополнительный файл",
}

QUESTION_FILE_KEYS = {
    "q12_diagnosis_file_id": "📎 Диагноз (файл)",
    "q17_confirmation_file": "📎 Подтверждающие документы (файл)",
    "q18_income_doc": "📎 Справка о доходах (файл)",
    "q19_children_docs": "📎 Документы на детей (файлы)",
    "q22_housing_doc": "📎 Документ на жильё (файл)",
    "q24_additional_file": "📎 Дополнительный файл",
}

def create_safe_filename(name: str) -> str:
    """Создает безопасное имя файла/папки, убирая недопустимые символы"""
    unsafe_chars = '<>:"/\\|?*'
    safe_name = name
    for char in unsafe_chars:
        safe_name = safe_name.replace(char, '_')
    return safe_name

def create_local_folder(name: str, parent_path: str = None) -> str:
    """Создает локальную папку и возвращает путь к ней"""
    if parent_path is None:
        parent_path = BASE_STORAGE_PATH
    
    # Создаем базовую папку если её нет
    Path(parent_path).mkdir(parents=True, exist_ok=True)
    
    # Создаем безопасное имя папки
    safe_name = create_safe_filename(name)
    folder_path = os.path.join(parent_path, safe_name)
    
    # Создаем папку
    Path(folder_path).mkdir(parents=True, exist_ok=True)
    
    return folder_path

def save_questionnaire_to_excel(user_data: dict, folder_path: str, filename: str = "Анкета.xlsx") -> str:
    """Сохраняет анкету в Excel файл с двумя столбцами: Вопрос и Ответ"""
    try:
        # Подготавливаем данные для таблицы
        questions = []
        answers = []
        
        # Добавляем ФИО первым
        full_name = user_data.get('q1_full_name', '')
        if full_name:
            questions.append("ФИО")
            answers.append(full_name)

        # Обрабатываем остальные вопросы
        for key, value in user_data.items():
            if key == "q1_full_name" or key == "full_name":
                continue
                
            label = QUESTION_LABELS.get(key, key)

            if key in QUESTION_FILE_KEYS:
                # Для файлов показываем статус наличия
                if isinstance(value, list):
                    answer = f"{len(value)} файл(ов) ✅" if value else "—"
                else:
                    answer = "Есть файл ✅" if value else "—"
            else:
                # Для обычных ответов
                answer = str(value) if value is not None else "—"
            
            questions.append(label)
            answers.append(answer)

        # Создаем DataFrame
        df = pd.DataFrame({
            'Вопрос': questions,
            'Ответ': answers
        })

        # Сохраняем в Excel
        excel_path = os.path.join(folder_path, filename)
        
        # Создаем Excel writer с настройками
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Анкета', index=False)
            
            # Получаем worksheet для настройки форматирования
            worksheet = writer.sheets['Анкета']
            
            # Настраиваем ширину столбцов
            worksheet.column_dimensions['A'].width = 45  # Столбец "Вопрос"
            worksheet.column_dimensions['B'].width = 65  # Столбец "Ответ"
            
            # Импортируем стили для форматирования
            from openpyxl.styles import Alignment, Border, Side, PatternFill, Font
            
            # Определяем стили
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")  # Синий фон для заголовков
            header_font = Font(color="FFFFFF", bold=True, size=12)  # Белый жирный текст
            
            question_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")  # Светло-серый для вопросов
            question_font = Font(bold=True, size=11)  # Жирный текст для вопросов
            
            answer_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")  # Белый для ответов
            answer_font = Font(size=11)  # Обычный текст для ответов
            
            # Создаем границы
            thin_border = Border(
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000'),
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000')
            )
            
            # Форматируем заголовки (первая строка)
            for col in range(1, 3):  # Столбцы A и B
                cell = worksheet.cell(row=1, column=col)
                cell.fill = header_fill
                cell.font = header_font
                cell.border = thin_border
                cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
            
            # Форматируем остальные строки
            for row in range(2, len(df) + 2):  # Начинаем со второй строки
                # Столбец A (Вопросы) - серый фон
                question_cell = worksheet.cell(row=row, column=1)
                question_cell.fill = question_fill
                question_cell.font = question_font
                question_cell.border = thin_border
                question_cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')
                
                # Столбец B (Ответы) - белый фон
                answer_cell = worksheet.cell(row=row, column=2)
                answer_cell.fill = answer_fill
                answer_cell.font = answer_font
                answer_cell.border = thin_border
                answer_cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')
            
            # Устанавливаем высоту строк для лучшего отображения
            for row in range(1, len(df) + 2):
                worksheet.row_dimensions[row].height = 25

        return excel_path

    except Exception as e:
        print(f"❌ Ошибка при сохранении Excel: {e}")
        raise

def save_text_to_local_file(content: str, filename: str, folder_path: str) -> str:
    """Сохраняет текст в локальный файл (оставлено для совместимости)"""
    safe_filename = create_safe_filename(filename)
    file_path = os.path.join(folder_path, safe_filename)
    
    with open(file_path, 'w', encoding='utf-8') as f:
        f.write(content)
    
    return file_path

async def save_telegram_file_locally(file_id: str, folder_path: str, filename: str, bot: Bot, user_id: int = None) -> str:
    """Скачивает файл из Telegram и сохраняет локально"""
    try:
        file = await bot.get_file(file_id)
        file_path = file.file_path
        file_bytes = await bot.download_file(file_path)
        
        # Получаем расширение файла
        ext = os.path.splitext(file_path)[-1] or ".bin"
        safe_filename = create_safe_filename(filename) + ext
        
        local_file_path = os.path.join(folder_path, safe_filename)
        
        # Сохраняем файл
        with open(local_file_path, 'wb') as f:
            f.write(file_bytes.read())
        
        log_user_action(
            user_id=user_id,
            action="File saved locally",
            state="save_telegram_file_locally",
            extra_data=f"File: {safe_filename}, Path: {local_file_path}"
        )
        
        return local_file_path

    except Exception as e:
        log_error(
            user_id=user_id,
            error=e,
            action="Failed to save file locally",
            state="save_telegram_file_locally",
        )
        print(f"❌ Error saving file {file_id}: {e}")
        raise

async def save_full_questionnaire_locally(user_data: dict, bot: Bot, user_id: int = None) -> str:
    """Сохраняет всю анкету в локальную папку с Excel файлом"""
    try:
        full_name = user_data.get('q1_full_name', 'Пациент')
        birth_date = user_data.get('q2_birth_date', 'Unknown')
        
        # Создаем основную папку для пациента
        folder_name = f"Анкета_пациента_{full_name}_{birth_date}"
        patient_folder_path = create_local_folder(folder_name)
        
        # Сохраняем анкету в Excel формате
        excel_file_path = save_questionnaire_to_excel(user_data, patient_folder_path)
        
        log_user_action(
            user_id=user_id,
            action="Questionnaire Excel saved locally",
            state="save_full_questionnaire_locally",
            extra_data=f"File: {excel_file_path}"
        )

        # Создаем общую папку для всех файлов пациента
        files_folder_path = create_local_folder("Файлы", parent_path=patient_folder_path)
        
        log_user_action(
            user_id=user_id,
            action="Created files folder for patient",
            state="save_full_questionnaire_locally",
            extra_data=f"Files folder: {files_folder_path}"
        )

        # Сохраняем файлы-вложения
        for field, folder_name in QUESTION_FILE_KEYS.items():
            file_value = user_data.get(field)
            if not file_value:
                continue

            # Создаем подпапку для каждого типа документов внутри папки "Файлы"
            subfolder_path = create_local_folder(folder_name, parent_path=files_folder_path)
            
            log_user_action(
                user_id=user_id,
                action="Created subfolder for attachments",
                state="save_full_questionnaire_locally",
                extra_data=f"{folder_name} (Field: {field})"
            )

            if isinstance(file_value, list):
                # Если несколько файлов
                for idx, file_id in enumerate(file_value):
                    if file_id:
                        await save_telegram_file_locally(
                            file_id, 
                            subfolder_path, 
                            f"{folder_name}_{idx+1}", 
                            bot, 
                            user_id
                        )
            else:
                # Если один файл
                await save_telegram_file_locally(
                    file_value, 
                    subfolder_path, 
                    folder_name, 
                    bot, 
                    user_id
                )

        log_user_action(
            user_id=user_id,
            action="Finished saving questionnaire locally",
            state="save_full_questionnaire_locally",
            extra_data=f"Patient folder: {patient_folder_path}"
        )
        
        return patient_folder_path

    except Exception as e:
        log_error(
            user_id=user_id,
            error=e,
            action="Error saving questionnaire locally",
            state="save_full_questionnaire_locally",
        )
        print(f"❌ Error in save_full_questionnaire_locally: {e}")
        raise

def delete_local_folder(folder_path: str) -> bool:
    """Удаляет локальную папку со всем содержимым"""
    try:
        import shutil
        if os.path.exists(folder_path):
            shutil.rmtree(folder_path)
            print(f"✅ Папка {folder_path} успешно удалена.")
            return True
        else:
            print(f"⚠️ Папка {folder_path} не найдена.")
            return False
    except Exception as e:
        print(f"❌ Ошибка при удалении папки: {e}")
        return False

def get_patient_folder_path(full_name: str, birth_date: str) -> str:
    """Возвращает путь к папке пациента"""
    folder_name = f"Анкета_пациента_{create_safe_filename(full_name)}_{birth_date}"
    return os.path.join(BASE_STORAGE_PATH, folder_name)

def list_all_questionnaires() -> list:
    """Возвращает список всех сохранённых анкет"""
    if not os.path.exists(BASE_STORAGE_PATH):
        return []
    
    questionnaires = []
    for item in os.listdir(BASE_STORAGE_PATH):
        item_path = os.path.join(BASE_STORAGE_PATH, item)
        if os.path.isdir(item_path):
            questionnaires.append({
                'name': item,
                'path': item_path,
                'created': os.path.getctime(item_path)
            })
    
    return sorted(questionnaires, key=lambda x: x['created'], reverse=True)